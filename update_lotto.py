import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os
import time

def update_excel():
    file_name = "lotto_prob.xlsx"
    if not os.path.exists(file_name):
        print("파일을 찾을 수 없습니다.")
        return

    try:
        # 수식 보존을 위해 data_only=False 유지
        wb = load_workbook(file_name, data_only=False)
        ws = wb['원본']
        
        # 1. 실제 데이터가 있는 마지막 행 찾기
        real_last_row = 1
        last_drw = 0
        for i in range(1, ws.max_row + 1):
            val = ws.cell(row=i, column=1).value
            if isinstance(val, int):
                real_last_row = i
                last_drw = val
        
        target_drw = last_drw + 1
        print(f"현재 엑셀 최신 회차: {last_drw}회")

        # 2. 네이버 크롤링
        url = f"https://search.naver.com/search.naver?query={target_drw}회로또"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"}
        
        res = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(res.text, 'html.parser')
        balls = soup.select('.ball')

        if balls and len(balls) >= 7:
            nums = [int(b.text) for b in balls[:7]]
            write_row = real_last_row + 1
            
            # A열: 회차, B~G열: 번호, H열: 보너스
            ws.cell(row=write_row, column=1, value=target_drw)
            for i, num in enumerate(nums):
                ws.cell(row=write_row, column=i + 2, value=num)
            
            # 파일 저장 (수식은 엑셀을 열 때 자동으로 재계산됨)
            wb.save(file_name)
            print(f"✅ {target_drw}회차 업데이트 성공!")
        else:
            print(f"ℹ️ {target_drw}회차는 아직 발표되지 않았습니다.")

    except Exception as e:
        print(f"❌ 에러 발생: {e}")

if __name__ == "__main__":
    update_excel()
